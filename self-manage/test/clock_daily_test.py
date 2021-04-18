import time

import wx
import os
from openpyxl import load_workbook


from views.Dialogs import NotExsit,AddSuccess
path1 = os.path.abspath('..')
excel_path = './data/excel/2020-11.xlsx'

sampleList = ['很好', '较好', '一般', '较差', '很差']


class ClockDaily(wx.Frame):
    def __init__(self, parent):
        super().__init__(parent)

        # 用来记录打卡项，也方便日后生成或者删除打卡项
        self.clock_list = [u'时间', u'早起', u'早睡', u'日语', u'英语', u'记账']
        self.clock_data = ['否' for i in range(len(self.clock_list))]

        self.InitUi()

    def InitUi(self):
        Box1 = wx.BoxSizer(wx.VERTICAL)
        panel = wx.Panel(self)

        self.SetBackgroundColour('white')
        Box1.AddSpacer(90)

        # 生成文本与单选框
        for i in range(0, len(self.clock_list)):
            self.ch = wx.CheckBox(self, label=self.clock_list[i], size=(90, 30))
            self.ch.SetValue(False)
            Box1.Add(self.ch, border=15)

        # 心情状况
        Box3 = wx.BoxSizer(wx.HORIZONTAL)
        self.mood1 = wx.StaticText(self, label=u"心情状况", size=(50, 20))
        Box3.Add(self.mood1, border=5)
        Box3.AddSpacer(20)
        self.ch1 = wx.Choice(self, -1, (60, 50), choices=sampleList, )
        # self.Bind(wx.EVT_CHOICE, self.EvtChoice1, self.ch)
        Box3.Add(self.ch1, flag=wx.EXPAND | wx.RIGHT, border=20)
        Box1.AddSpacer(10)
        Box1.Add(Box3, border=20)

        # 打卡项，身体状况
        Box4 = wx.BoxSizer(wx.HORIZONTAL)
        self.body1 = wx.StaticText(self, label=u"身体状况", size=(50, 20))

        Box4.Add(self.body1, border=15)
        Box4.AddSpacer(10)
        self.ch2 = wx.Choice(self, -1, (60, 50), choices=sampleList)
        # self.Bind(wx.EVT_CHOICE, self.EvtChoice2, self.ch1)
        Box4.Add(self.ch2, flag=wx.EXPAND | wx.RIGHT, border=20)

        Box1.AddSpacer(20)
        Box1.Add(Box4, border=20)

        # 用来填充

        Box1.Add((-1, 10))
        # 提交按钮
        self.summit = wx.Button(self, label=u"提交", size=(90, 30))
        self.Bind(wx.EVT_BUTTON, self.OnClick, self.summit)

        Box1.Add(self.summit, border=15)

        Box2 = wx.BoxSizer(wx.VERTICAL)
        str1 = wx.StaticText(panel, label=u"    ", size=(10, 10))
        Box2.Add(str1)

        Box = wx.BoxSizer(wx.HORIZONTAL)
        Box.Add(Box2)
        Box.Add(Box1)

        self.SetSizer(Box)

    def OnClick(self, e):
        print(self.clock_data)
        if e.GetEventObject() == self.summit:
            for i in range(0,len(self.clock_list)):
                if self.ch.GetValue() :
                    self.clock_data[i] = '是'

        print (self.clock_data)


        try:

            wb = load_workbook(excel_path)
            ws = wb['每日打卡']
            print("文件打开成功")
            c = ws['A1']
            b = int(c.value)  # excel 单元格第一格用来记录上次写入的位置
            a = ws['A' + str(b)]
            print('单元格获取成功')
            a = ws.cell(row=b, column=1)

        except:
            dlg = NotExsit(None, -1)
            dlg.ShowModal()
            dlg.Destroy()

        # 用来防止误操作导致b的值与实际值不相符，再做一次判断
        while a.value:
            b += 1
            a = ws['A' + str(b)]
            if b >= 1500:
                break
        # 写入数据
        try:
            print('写数据')
            ws['A' + str(b)] = time.strftime('%Y-%m-%d', time.localtime())
            print('1')
            ws['B' + str(b)] = self.is_getup
            print('2')
            ws['D' + str(b)] = self.is_Japanese
            print('3')
            ws['E' + str(b)] = self.is_English
            print('4')
            ws['F' + str(b)] = self.is_study
            print('5')
            ws['H' + str(b)] = self.is_reading
            print('6')
            ws['I' + str(b)] = self.is_sleep
            print('写入数据')
            ws['A' + 1] = b + 1
            wb.save(excel_path)
            print('写入成功')
            dlg = AddSuccess(None, -1)
            print('对话框')
            dlg.ShowModal()
            dlg.Destroy()

        except:
            dlg = NotExsit(None, -1)
            dlg.ShowModal()
            dlg.Destroy()



def main():
    app = wx.App()
    lo= ClockDaily(None)
    lo.Show()
    app.MainLoop()

if __name__=='__main__':
    main()