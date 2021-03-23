import wx
from generic_bitmap_button import GenericBitmapButton
import wx.adv
import datetime
import time
from my_Validator import MyNumberValidator
from openpyxl import load_workbook
from wx import NewIdRef
import wx.lib.buttons as buttons
import images
from wx import NewId

expend = ['教育','餐饮','理财','日用','零食','交通','服饰美容','数码','住房','医疗']
ID_00 = NewId()
ID_01 = NewId()
ID_02 = NewId()
ID_03 = NewId()
ID_04 = NewId()
ID_05 = NewId()
ID_06 = NewId()
ID_07 = NewId()
ID_08 = NewId()
ID_09 = NewId()
class Expenditure(wx.Panel):
    def __init__(self,parent):
        super().__init__(parent)

        self.n_time = time.strftime('%Y-%m-%d', time.localtime())
        self.n_kind = expend[0]
        self.n_acount = 0
        self.n_remark = '无'
        self.initUi()
        self.Center()

    def initUi(self):
        self.SetBackgroundColour("white")

        box1 = wx.GridSizer(4,5,5,10)
        self.t_0= wx.StaticText(self,label = expend[0])
        self.bt_0 = buttons.GenBitmapToggleButton(self, ID_00, None)
        #self.Bind(wx.EVT_BUTTON, self.OnToggleButton, b)
        bmp = images._10.GetBitmap()
        mask = wx.Mask(bmp, wx.BLUE)
        bmp.SetMask(mask)
        self.bt_0.SetBitmapLabel(bmp)
        bmp = images._100.GetBitmap()
        mask = wx.Mask(bmp, wx.BLUE)
        bmp.SetMask(mask)
        self.bt_0.SetBitmapSelected(bmp)
        #self.bt_0.SetValue(True)
        self.bt_0.SetToggle(False)
        self.bt_0.SetInitialSize()
        self.bt_0.Bind(wx.EVT_BUTTON,self.KindSelect)

        self.t_1 = wx.StaticText(self, label=expend[1])
        self.bt_1 = buttons.GenBitmapToggleButton(self, ID_01, None)
        bmp = images._11.GetBitmap()
        mask = wx.Mask(bmp, wx.BLUE)
        bmp.SetMask(mask)
        self.bt_1.SetBitmapLabel(bmp)
        bmp = images._110.GetBitmap()
        mask = wx.Mask(bmp, wx.BLUE)
        bmp.SetMask(mask)
        self.bt_1.SetBitmapSelected(bmp)
        #self.bt_1.SetValue(True)
        self.bt_1.SetToggle(False)
        self.bt_1.SetInitialSize()
        self.bt_1.Bind(wx.EVT_BUTTON, self.KindSelect)

        self.t_2 = wx.StaticText(self, label=expend[2])
        self.bt_2 = buttons.GenBitmapToggleButton(self, ID_02, None)
        bmp = images._12.GetBitmap()
        mask = wx.Mask(bmp, wx.BLUE)
        bmp.SetMask(mask)
        self.bt_2.SetBitmapLabel(bmp)
        bmp = images._120.GetBitmap()
        mask = wx.Mask(bmp, wx.BLUE)
        bmp.SetMask(mask)
        self.bt_2.SetBitmapSelected(bmp)
        # self.bt_1.SetValue(True)
        self.bt_2.SetToggle(False)
        self.bt_2.SetInitialSize()
        self.bt_2.Bind(wx.EVT_BUTTON, self.KindSelect)

        self.t_3 = wx.StaticText(self, label=expend[3])
        self.bt_3 = buttons.GenBitmapToggleButton(self, ID_03, None)
        bmp = images._13.GetBitmap()
        mask = wx.Mask(bmp, wx.BLUE)
        bmp.SetMask(mask)
        self.bt_3.SetBitmapLabel(bmp)
        bmp = images._130.GetBitmap()
        mask = wx.Mask(bmp, wx.BLUE)
        bmp.SetMask(mask)
        self.bt_3.SetBitmapSelected(bmp)
        # self.bt_1.SetValue(True)
        self.bt_3.SetToggle(False)
        self.bt_3.SetInitialSize()
        self.bt_3.Bind(wx.EVT_BUTTON, self.KindSelect)

        self.t_4 = wx.StaticText(self, label=expend[4])
        self.bt_4 = buttons.GenBitmapToggleButton(self, ID_04, None)
        bmp = images._14.GetBitmap()
        mask = wx.Mask(bmp, wx.BLUE)
        bmp.SetMask(mask)
        self.bt_4.SetBitmapLabel(bmp)
        bmp = images._140.GetBitmap()
        mask = wx.Mask(bmp, wx.BLUE)
        bmp.SetMask(mask)
        self.bt_4.SetBitmapSelected(bmp)
        # self.bt_1.SetValue(True)
        self.bt_4.SetToggle(False)
        self.bt_4.SetInitialSize()
        self.bt_4.Bind(wx.EVT_BUTTON, self.KindSelect)

        self.t_5 = wx.StaticText(self, label=expend[5])
        self.bt_5 = buttons.GenBitmapToggleButton(self, ID_05, None)
        bmp = images._15.GetBitmap()
        mask = wx.Mask(bmp, wx.BLUE)
        bmp.SetMask(mask)
        self.bt_5.SetBitmapLabel(bmp)
        bmp = images._150.GetBitmap()
        mask = wx.Mask(bmp, wx.BLUE)
        bmp.SetMask(mask)
        self.bt_5.SetBitmapSelected(bmp)
        # self.bt_1.SetValue(True)
        self.bt_5.SetToggle(False)
        self.bt_5.SetInitialSize()
        self.bt_5.Bind(wx.EVT_BUTTON, self.KindSelect)

        self.t_6 = wx.StaticText(self, label=expend[6])
        self.bt_6 = buttons.GenBitmapToggleButton(self, ID_06, None)
        bmp = images._16.GetBitmap()
        mask = wx.Mask(bmp, wx.BLUE)
        bmp.SetMask(mask)
        self.bt_6.SetBitmapLabel(bmp)
        bmp = images._160.GetBitmap()
        mask = wx.Mask(bmp, wx.BLUE)
        bmp.SetMask(mask)
        self.bt_6.SetBitmapSelected(bmp)
        # self.bt_1.SetValue(True)
        self.bt_6.SetToggle(False)
        self.bt_6.SetInitialSize()
        self.bt_6.Bind(wx.EVT_BUTTON, self.KindSelect)

        self.t_7 = wx.StaticText(self, label=expend[7])
        self.bt_7 = buttons.GenBitmapToggleButton(self, ID_07, None)
        bmp = images._17.GetBitmap()
        mask = wx.Mask(bmp, wx.BLUE)
        bmp.SetMask(mask)
        self.bt_7.SetBitmapLabel(bmp)
        bmp = images._170.GetBitmap()
        mask = wx.Mask(bmp, wx.BLUE)
        bmp.SetMask(mask)
        self.bt_7.SetBitmapSelected(bmp)
        # self.bt_1.SetValue(True)
        self.bt_7.SetToggle(False)
        self.bt_7.SetInitialSize()
        self.bt_7.Bind(wx.EVT_BUTTON, self.KindSelect)

        self.t_8 = wx.StaticText(self, label=expend[8])
        self.bt_8 = buttons.GenBitmapToggleButton(self, ID_08, None)
        bmp = images._18.GetBitmap()
        mask = wx.Mask(bmp, wx.BLUE)
        bmp.SetMask(mask)
        self.bt_8.SetBitmapLabel(bmp)
        bmp = images._180.GetBitmap()
        mask = wx.Mask(bmp, wx.BLUE)
        bmp.SetMask(mask)
        self.bt_8.SetBitmapSelected(bmp)
        # self.bt_1.SetValue(True)
        self.bt_8.SetToggle(False)
        self.bt_8.SetInitialSize()
        self.bt_8.Bind(wx.EVT_BUTTON, self.KindSelect)

        self.t_9 = wx.StaticText(self, label=expend[9])
        self.bt_9 = buttons.GenBitmapToggleButton(self, ID_09, None)
        bmp = images._19.GetBitmap()
        mask = wx.Mask(bmp, wx.BLUE)
        bmp.SetMask(mask)
        self.bt_9.SetBitmapLabel(bmp)
        bmp = images._190.GetBitmap()
        mask = wx.Mask(bmp, wx.BLUE)
        bmp.SetMask(mask)
        self.bt_9.SetBitmapSelected(bmp)
        # self.bt_1.SetValue(True)
        self.bt_9.SetToggle(False)
        self.bt_9.SetInitialSize()
        self.bt_9.Bind(wx.EVT_BUTTON, self.KindSelect)

        box1.AddMany([(self.bt_0,0,wx.EXPAND),(self.bt_1,0,wx.EXPAND),
                      (self.bt_2,0,wx.EXPAND),(self.bt_3,0,wx.EXPAND),
                      (self.bt_4,0,wx.EXPAND),
                      (self.t_0, 0, wx.LEFT),
                      (self.t_1, 0, wx.LEFT),(self.t_2, 0, wx.LEFT),
                      (self.t_3, 0, wx.LEFT),(self.t_4, 0, wx.LEFT),

                      (self.bt_5,0,wx.EXPAND),(self.bt_6,0,wx.EXPAND),
                      (self.bt_7,0,wx.EXPAND),(self.bt_8,0,wx.EXPAND),
                      (self.bt_9,0,wx.EXPAND),
                      (self.t_5, 0, wx.EXPAND),(self.t_6, 0, wx.EXPAND),
                      (self.t_7, 0, wx.EXPAND),
                      (self.t_8, 0, wx.EXPAND), (self.t_9, 0, wx.EXPAND),
                      ])

        box2 = wx.BoxSizer(wx.HORIZONTAL)

        self.time1 = wx.adv.DatePickerCtrl(self,style = wx.adv.DP_DROPDOWN
                                      | wx.adv.DP_SHOWCENTURY
                                      | wx.adv.DP_ALLOWNONE)

        self.Bind(wx.adv.EVT_DATE_CHANGED,self.OnDateSelect,self.time1)

        self.t_acount = wx.TextCtrl(self, -1,"金额",validator = MyNumberValidator())
        self.Bind(wx.EVT_TEXT,self.EvtText1,self.t_acount)

        box2.Add(self.time1,flag=wx.LEFT,border=10)
        box2.Add(self.t_acount,flag = wx.RIGHT|wx.EXPAND,border =10)

        self.t_remarks = wx.TextCtrl(self,-1, "备注",size = (50,100))
        self.Bind(wx.EVT_TEXT, self.EvtText2, self.t_remarks)

        self.bt_commit = GenericBitmapButton(self, 'commit', -1)
        self.bt_commit.SetToolTip('提交')
        self.Bind(wx.EVT_BUTTON,self.Commit,self.bt_commit)

        box3 = wx.BoxSizer(wx.VERTICAL)

        box3.AddSpacer(10)
        box3.Add(box1,flag=wx.ALIGN_CENTER)
        box3.Add(box2,flag = wx.ALIGN_CENTER)
        box3.Add(self.t_remarks,flag=wx.EXPAND|wx.ALL,border=20)
        box4= wx.BoxSizer(wx.HORIZONTAL)
        box4.AddSpacer(280)
        box4.Add(self.bt_commit,flag = wx.RIGHT,border = 10)
        box3.Add(box4,flag = wx.RIGHT,border = 10)

        self.SetSizer(box3)

    def EvtText1(self,e):
        self.n_acount= str(e.GetString())

    def EvtText2(self,e):
        self.n_remark = e.GetString()

    def OnDateSelect(self,e):
        #print(str(e.GetString()))
        self.n_time = str(e.GetDate())[0:10]
        print(self.n_time)

    def KindSelect(self,e):
        self.bt_0.SetToggle(False)
        self.bt_1.SetToggle(False)
        self.bt_2.SetToggle(False)
        self.bt_3.SetToggle(False)
        self.bt_4.SetToggle(False)
        self.bt_5.SetToggle(False)
        self.bt_6.SetToggle(False)
        self.bt_7.SetToggle(False)
        self.bt_8.SetToggle(False)
        self.bt_9.SetToggle(False)
        #print(e.GetId)
        #print(ID_03)
        #print(type(e.GetId))
        #print(type(ID_03))
        if e.GetId() == ID_00:
            self.bt_0.SetToggle(True)
            self.n_kind = expend[0]
        if e.GetId() == ID_01:
            self.bt_1.SetToggle(True)
            self.n_kind =expend[1]
        if e.GetId() == ID_02:
            self.bt_2.SetToggle(True)
            self.n_kind =expend[2]
        if e.GetId() == ID_03:
            self.bt_3.SetToggle(True)
            self.n_kind =expend[3]
            #print('good')
        if e.GetId() == ID_04:
            self.bt_4.SetToggle(True)
            self.n_kind =expend[4]
        if e.GetId() == ID_05:
            self.bt_5.SetToggle(True)
            self.n_kind =expend[5]
        if e.GetId() == ID_06:
            self.bt_6.SetToggle(True)
            self.n_kind =expend[6]
        if e.GetId() == ID_07:
            self.bt_7.SetToggle(True)
            self.n_kind =expend[7]
        if e.GetId() == ID_08:
            self.bt_8.SetToggle(True)
            self.n_kind =expend[8]
        if e.GetId() == ID_09:
            self.bt_9.SetToggle(True)
            self.n_kind =expend[9]


        #print(self.n_kind)

    def Commit(self,e):

        print(self.n_time)
        print('1')
        #print(self.n_acount)

        #print(self.n_remark)
        #print(self.n_kind)

        try:
            wb=load_workbook('./data/account.xlsx')
            ws=wb['支出']
            #print("文件打开成功")
            #i_row = 1
            #i_column = 1
            c=ws['A1']
            b=int(c.value)    #excel 单元格第一格用来记录上次写入的位置
            a=ws.cell(row = b,column = 1)
            #print(str(b.value))
            #print('单元格获取成功')
            #c=ws['A1']
            #print(c.value)

        except:
            dlg = wx.MessageDialog(self, '文件打开失败',
                                   '失败',
                                   wx.OK | wx.ICON_INFORMATION
                                   # wx.YES_NO | wx.NO_DEFAULT | wx.CANCEL | wx.ICON_INFORMATION
                                   )
            dlg.ShowModal()
            dlg.Destroy()
            
        while a.value:
            #使用遍历的方式来对单元格进行检查
            #如果获取的单元格的内容为空，则表示可以写入
            #b参数为行数
            #print(b)
            #print(type(a.value))
            #print (len(a))
            

            b+=1
            #i_column +=1
            #print('行数是'+str(b))
            a=ws['A'+str(b)]
            #print(a.value)
            
            if b>=1500:
                
                break
                #a = ws.cell(row = i_row,column = i_column)

        if self.n_kind and self.n_time and self.n_acount:

            try:

                ws['A'+str(b)]=self.n_time
                #print('1')
                ws['B'+str(b)]=self.n_kind
                #print('2')
                ws['C'+str(b)]=self.n_acount
                #print('3')
                ws['D'+str(b)]=self.n_remark
                #print('4')
                ws['A1'] = b+1

                wb.save('./data/account.xlsx')

            except:
                dlg = wx.MessageDialog(self, '数据写入失败',
                                       '失败',
                                       wx.OK | wx.ICON_INFORMATION
                                       # wx.YES_NO | wx.NO_DEFAULT | wx.CANCEL | wx.ICON_INFORMATION
                                       )
                dlg.ShowModal()
                dlg.Destroy()
            #print('文件写入成功')

        else:
            dlg = wx.MessageDialog(self, '请填写完整数据',
                                   '失败',
                                   wx.OK | wx.ICON_INFORMATION
                                   # wx.YES_NO | wx.NO_DEFAULT | wx.CANCEL | wx.ICON_INFORMATION
                                   )
            dlg.ShowModal()
            dlg.Destroy()

        self.bt_commit.Enable(False)
        time.sleep(1)
        self.bt_commit.Enable(True)

        dlg = wx.MessageDialog(self, '写入成功',
                               '成功',
                               wx.OK | wx.ICON_INFORMATION
                               # wx.YES_NO | wx.NO_DEFAULT | wx.CANCEL | wx.ICON_INFORMATION
                               )
        dlg.ShowModal()
        dlg.Destroy()




def main():

    app = wx.App()
    ex = Expenditure()
    ex.Show()
    app.MainLoop()


if __name__ == '__main__':
    main()





class Read_Money(wx.Frame):
    def __init__(self,parent):
        super().__init__(parent,title=u'记账',
                         style=wx.MINIMIZE|wx.SYSTEM_MENU|wx.CLOSE_BOX|wx.CAPTION)

        self.initUi()
        self.Center()

    def initUi(self):
        self.SetBackgroundColour("white")


