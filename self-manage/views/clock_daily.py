import time

import wx
import os
from openpyxl import load_workbook
from .Dialogs import NotExsit, AddSuccess,HavedExit,BDOpenFail
import sqlite3

path1 = os.path.abspath('..')
excel_path = './data/2020-11.xlsx'

sampleList = ['很好', '较好', '一般', '较差', '很差']

class ClockDaily(wx.Panel):
    def __init__(self, parent):
        super().__init__(parent)
        self.is_getup = '否'
        self.is_sleep = '否'
        self.is_Japanese = '否'
        self.is_English = '否'
        self.is_reading = '否'
        self.is_study = '否'
        self.is_account = '否'
        self.mood = '一般'
        self.body = '一般'
        self.InitUi()

    def InitUi(self):
        Box1=wx.BoxSizer(wx.VERTICAL)
        panel = wx.Panel(self)

        self.SetBackgroundColour('white')
        Box1.AddSpacer(90)

        #是否早起
        self.get_up = wx.CheckBox(self, label=u"早起", size=(90, 30))
        self.get_up.SetValue(False)  # 设置当前是否被选中
        Box1.Add(self.get_up, border=15)

        # 是否早睡
        self.sleep = wx.CheckBox(self, label=u"早睡", size=(90, 30))
        self.sleep.SetValue(False)  # 设置当前是否被选中
        Box1.Add(self.sleep, border=15)

        # 是否学习了日语
        self.Japanese = wx.CheckBox(self, label=u"日语学习", size=(90, 30))
        self.Japanese.SetValue(False)  # 设置当前是否被选中
        Box1.Add(self.Japanese, border=15)

        # 是否背诵了英语单词
        self.English = wx.CheckBox(self, label=u"英语学习", size=(90, 30))
        self.English.SetValue(False)  # 设置当前是否被选中
        Box1.Add(self.English, border=15)

        # 是否进行了学习
        self.study = wx.CheckBox(self, label=u"学习", size=(90, 30))
        self.study.SetValue(False)  # 设置当前是否被选中
        Box1.Add(self.study, border=15)

        # 是否有做阅读
        self.reading = wx.CheckBox(self, label=u"阅读", size=(90, 30))
        self.reading.SetValue(False)  # 设置当前是否被选中
        Box1.Add(self.reading, border=15)

        self.account = wx.CheckBox(self, label=u"记账", size=(90, 30))
        self.reading.SetValue(False)  # 设置当前是否被选中
        Box1.Add(self.account, border=15)

        # 心情状况
        Box3 = wx.BoxSizer(wx.HORIZONTAL)
        self.mood1 = wx.StaticText(self, label=u"心情状况", size=(50, 20))
        Box3.Add(self.mood1, border=5)
        Box3.AddSpacer(20)
        self.ch = wx.Choice(self, -1, (60, 50), choices=sampleList, )
        self.Bind(wx.EVT_CHOICE,self.EvtChoice0,self.ch)
        Box3.Add(self.ch, flag=wx.EXPAND | wx.RIGHT, border=20)
        Box1.AddSpacer(10)
        Box1.Add(Box3, border=20)

        # 打卡项，身体状况
        Box4 = wx.BoxSizer(wx.HORIZONTAL)
        self.body1 = wx.StaticText(self, label=u"身体状况", size=(50, 20))
        Box4.Add(self.body1, border=15)
        Box4.AddSpacer(20)
        self.ch1 = wx.Choice(self, -1, (60, 50), choices=sampleList)
        self.Bind(wx.EVT_CHOICE, self.EvtChoice1, self.ch1)
        Box4.Add(self.ch1, flag=wx.EXPAND | wx.RIGHT, border=20)

        Box1.AddSpacer(20)
        Box1.Add(Box4, border=20)

        # 用来填充

        Box1.Add((-1, 10))

        # 提交按钮
        self.summit = wx.Button(self, label=u"提交", size=(90, 30))
        self.Bind(wx.EVT_BUTTON, self.OnClick, self.summit)
        Box1.Add(self.summit, border=15)

        Box2 = wx.BoxSizer(wx.VERTICAL)
        str1 = wx.StaticText(panel, label=u"    ", size=(10, 10))
        Box2.Add(str1)

        Box = wx.BoxSizer(wx.HORIZONTAL)
        Box.Add(Box2)
        Box.Add(Box1)

        self.SetSizer(Box)

    def EvtChoice0(self,e):
        self.mood = e.GetString()

    def EvtChoice1(self,e):
        self.body = e.GetString()

    def OnClick(self, e):
        if e.GetEventObject() == self.summit:
            if self.get_up.GetValue():
                self.is_getup = '是'
            if self.sleep.GetValue():
                self.is_sleep = '是'
            if self.Japanese.GetValue():
                self.is_Japanese = '是'
            if self.English.GetValue():
                self.is_English = '是'
            if self.reading.GetValue():
                self.is_reading = '是'
            if self.study.GetValue():
                self.is_study = '是'
            if self.account.GetValue():
                self.is_account = '是'

        self.excel_save()
        self.db_save()

        #self.db_save()

    def excel_save(self):
        try:

            wb=load_workbook(excel_path)
            ws=wb['每日打卡']
            b = int(ws['A1'].value)  # excel 单元格第一格用来记录上次写入的位置
            a = ws.cell(row=b, column=1)

            # 用来防止误操作导致b的值与实际值不相符，再做一次判断
            while a.value:
                b += 1
                a = ws['A' + str(b)]
                if b >= 1500:
                    break

        except:
            dlg = NotExsit(None, -1)
            dlg.ShowModal()
            dlg.Destroy()


        #写入数据
        try:

            #判断今天是否已经打过卡
            if ws['A'+str(b-1)].value == time.strftime('%Y-%m-%d', time.localtime()):
                dlg = HavedExit(None, -1)
                dlg.ShowModal()
                dlg.Destroy()
            else:
                b = str(b)
                ws['A'+b]=time.strftime('%Y-%m-%d', time.localtime())
                ws['B'+b]= self.is_getup
                ws['C'+b]= self.is_sleep
                ws['D'+b]= self.is_Japanese
                ws['E'+b]=self.is_English
                ws['F'+b]=self.is_study
                ws['G'+b]=self.is_reading
                ws['H'+b] = self.is_account
                ws['I'+b] = self.mood
                ws['J'+b] = self.body


                ws['A'+str(1)] = int(b)+1
                wb.save(excel_path)

                dlg = AddSuccess(None,-1)
                dlg.ShowModal()
                dlg.Destroy()

        except:
            dlg=NotExsit(None,-1)
            dlg.ShowModal()
            dlg.Destroy()

    def db_save(self):
        #将数据保存在数据库中
        #这里暂时用sqlite3来保存信息

        time1 = time.strftime('%Y-%m-%d', time.localtime())
        try:
            #尝试连接数据库
            conn = sqlite3.connect('my_record.db')

            cursor = conn.cursor()     #创建游标

            #判断表是否存在，若不存在，则新建该表
            cursor.execute(
            '''CREATE TABLE IF NOT EXISTS clock_daily
            (ID INT PRIMARY KEY     ,
            x_time TEXT NOT NULL,
            get_up TEXT MOT NULL ,
            sleep TEXT MOT NULL, 
            Japanese TEXT MOT NULL,
            English TEXT MOT NULL,
            study TEXT MOT NULL,
            reading TEXT MOT NULL,
            account TEXT MOT NULL,
            mood TEXT MOT NULL,
            body TEXT MOT NULL)'''
            )


            #判断是否存在该条数据，如果存在，则弹出提示说今天已打卡
            c = cursor.execute("SELECT x_time FROM clock_daily")
            is_exist = False

            for row in c :

                if row[0] == time1 :
                    is_exist = True
                    break


            if is_exist:
                # 表示存在该条数据
                dlg = HavedExit(None, -1)
                dlg.ShowModal()
                dlg.Destroy()

            else:
                cursor.execute('INSERT INTO clock_daily(x_time,get_up,sleep,Japanese,English,study,reading,account,mood,body)'
                           ' VALUES (?,?,?,?,?,?,?,?,?,?)',(time1,self.is_getup,self.is_sleep
                            ,self.is_Japanese,self.is_English,self.is_study,self.is_reading,self.is_account,self.mood,self.body))

                conn.commit()
                conn.close()

        except:
            dlg = BDOpenFail(None, -1)
            dlg.ShowModal()
            dlg.Destroy()



