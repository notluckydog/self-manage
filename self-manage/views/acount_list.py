import wx
import wx.grid as grid
from openpyxl import load_workbook

class Data_list(grid.Grid):
    def __init__(self):
        grid.Grid.__init__(self)

        self.CreateGrid(20,4)


        self.x_time = []
        self.x_kind = []
        self.x_account = []
        self.x_remark = []

        try:
            wb = load_workbook('./data/account.xlsx')
            ws = wb['支出']
            #c = ws['A1']
            b = int(ws['A1'].value)  # excel 单元格第一格用来记录上次写入的位置
            a = ws.cell(row =2,column = b)

            if b<=23:
                d = b-3
                for clo in range(0,d):
                    time1 = str(ws.cell(row=0, column=clo).value)
                    self.x_time.append(time1)
                    kind1  =str(ws.cell(row=1, column=clo).value)
                    self.x_kind.append(kind1)
                    acount = str(ws.cell(row=2, column=clo).value)
                    self.x_account.append(acount)
                    remark = str(ws.cell(row=3, column=clo).value)
                    self.x_remark.append(remark)
                    b-=1


            else:
                for clo in range(0, 20):
                    time1 = str(ws.cell(row=0, column=b).vaule)
                    self.x_time.append(time1)
                    kind1 = str(ws.cell(row=1, column=b).vaule)
                    self.x_kind.append(kind1)
                    acount = str(ws.cell(row=2, column=b).vaule)
                    self.x_account.append(acount)
                    remark = str(ws.cell(row=3, column=b).vaule)
                    self.x_remark.append(remark)
                    b -= 1

            for i in range(0,len(self.x_time)):
                self.SetCellValue(i,0,self.x_time)
                self.SetCellValue(i, 0, self.x_kind)
                self.SetCellValue(i, 0, self.x_account)
                self.SetCellValue(i, 0, self.x_remark)

            self.SetColSize(0, 150)
            self.SetColSize(5, 150)

        except:
            dlg = wx.MessageDialog(self, '数据读取失败',
                                   '失败',
                                   wx.OK | wx.ICON_INFORMATION
                                   # wx.YES_NO | wx.NO_DEFAULT | wx.CANCEL | wx.ICON_INFORMATION
                                   )
            dlg.ShowModal()
            dlg.Destroy()

