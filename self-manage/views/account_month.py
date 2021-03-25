import wx
from matplotlib.backends import backend_wxagg
from matplotlib.figure import Figure
import matplotlib
from matplotlib.backends.backend_wxagg import FigureCanvasWxAgg as FigureCanvas
import wx.adv
import time
from openpyxl import load_workbook

imcome = ['薪资','退款','理财','兼职','还钱','借入','意外所得','报销','投资','其他']
expend = ['教育','餐饮','理财','日用','零食','交通','服饰美容','数码','住房','医疗']
year_list = ['2020','2021','2022']
month_list = ['一月','二月','三月','四月','五月','六月','七月','八月','九月','十月','十一月','十二月']


class AccountMonth(wx.Panel):
    def __init__(self,parent):
        super().__init__(parent)
        self.time_todey = time.strftime('%Y-%m-%d', time.localtime())
        self.X_year = self.time_todey[0:4]
        self.X_month = self.time_todey[5:7]
        if self.X_month in ['10','11','12']:
            pass
        else:
            self.X_month = self.X_month[1:2]
        self.i_list = [1]*10     #用来记录收入列表
        self.e_list = [1]*10     #用来记录支出列表
        self.i_space_list = [0.1]*10   #用来记录收入相关占比
        self.e_space_list = [0.1]*10   #用来记录支出相关占比
        self.i_count = 1
        self.e_count = 1
        self.Center()
        self.initUi()

    def initUi(self):
        self.SetBackgroundColour('white')

        # 准备数据
        self.GetIncome()
        self.GetExpendture()
        values = [26, 17, 21, 29, 7, 10]
        spaces = [0.05, 0.01, 0.01, 0.01, 0.01, 0.01]
        labels = ['Python', 'JavaScript',
                  'C++', 'Java', 'PHP', 'C']
        colors = ['dodgerblue', 'orangered','green','lime','yellow',
                  'limegreen', 'violet', 'gold', 'r','blue',]

        self.f = Figure(figsize=(5,8),dpi=100,tight_layout=True
                        )
        matplotlib.rcParams['font.sans-serif'] = ['SimHei']  # 用黑体显示中文
        matplotlib.rcParams['axes.unicode_minus'] = False

        self.sub = self.f.add_subplot(2,2,1,title ="支出明细")

        self.sub.pie(self.e_list, self.e_space_list, expend, colors,
               "%.1f%%", shadow=True, labeldistance=1.1,
                     radius=2,startangle=90)
        self.sub.axis("equal")

        self.sub = self.f.add_subplot(2,2,2,title = "收入明细")
        self.sub.pie(self.i_list, self.i_space_list, imcome, colors,
               "%.1f%%", shadow=True, labeldistance=1.1,
                     radius=2,startangle=90)

        self.sub.axis("equal")

        canvas = FigureCanvas(self, -1, self.f)
        canvas.draw()

        year_list = ['2020', '2021', '2022']
        self.ch_year = wx.Choice(self,-1,(85,18),choices = year_list)
        self.Bind(wx.EVT_CHOICE, self.OnChoices1, self.ch_year)
        self.t_year = wx.StaticText(self,label=u'年')

        month_list = ['1', '2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '12']
        self.ch_month = wx.Choice(self,-1,(85,18),choices = month_list)
        self.Bind(wx.EVT_CHOICE,self.OnChoices2,self.ch_month)
        self.t_month = wx.StaticText(self,label=u'月')

        box2 = wx.BoxSizer(wx.VERTICAL)
        self.t_0  = wx.StaticText(self,-1,label = u'消费总额'+str(self.e_count))
        self.t_00 = wx.StaticText(self,-1,label = expend[0]+': '+ str(self.e_list[0]))
        self.t_01 = wx.StaticText(self,-1,label = expend[1]+': '+ str(self.e_list[1]))
        self.t_02 = wx.StaticText(self,-1,label = expend[2]+': '+ str(self.e_list[2]))
        self.t_03 = wx.StaticText(self, -1, label=expend[3] + ': ' + str(self.e_list[3]))
        self.t_04 = wx.StaticText(self, -1, label=expend[4] + ': ' + str(self.e_list[4]))
        self.t_05 = wx.StaticText(self, -1, label=expend[5] + ': ' + str(self.e_list[5]))
        self.t_06 = wx.StaticText(self, -1, label=expend[6] + ': ' + str(self.e_list[6]))
        self.t_07 = wx.StaticText(self, -1, label=expend[7] + ': ' + str(self.e_list[7]))
        self.t_08 = wx.StaticText(self, -1, label=expend[8] + ': ' + str(self.e_list[8]))
        self.t_09 = wx.StaticText(self, -1, label=expend[9] + ': ' + str(self.e_list[9]))

        box2.Add(self.t_0, flag = wx.ALIGN_CENTER,border = 1)
        box2.Add(self.t_00,flag = wx.ALIGN_CENTER,border =1)
        box2.Add(self.t_01, flag=wx.ALIGN_CENTER, border=1)
        box2.Add(self.t_02, flag=wx.ALIGN_CENTER, border=1)
        box2.Add(self.t_03, flag=wx.ALIGN_CENTER, border=1)
        box2.Add(self.t_04, flag=wx.ALIGN_CENTER, border=1)
        box2.Add(self.t_05, flag=wx.ALIGN_CENTER, border=1)
        box2.Add(self.t_06, flag=wx.ALIGN_CENTER, border=1)
        box2.Add(self.t_07, flag=wx.ALIGN_CENTER, border=1)
        box2.Add(self.t_08, flag=wx.ALIGN_CENTER, border=1)
        box2.Add(self.t_09, flag=wx.ALIGN_CENTER, border=1)

        box3 = wx.FlexGridSizer(11,2,5,5)#用来展示总支出，每项支出
        self.t_ex = wx.StaticText(self,label=u'总支出')
        self.t_ex_0 = wx.StaticText(self,label = str(self.e_count))

        self.t_ex_00 = wx.StaticText(self,label =expend[0])
        self.t_ex_01 = wx.StaticText(self,label = str(self.e_list[0]))
        self.t_ex_10 = wx.StaticText(self, label=expend[1])
        self.t_ex_11 = wx.StaticText(self, label=str(self.e_list[1]))

        self.t_ex_20 = wx.StaticText(self, label=expend[2])
        self.t_ex_21 = wx.StaticText(self, label=str(self.e_list[2]))

        self.t_ex_30 = wx.StaticText(self, label=expend[3])
        self.t_ex_31 = wx.StaticText(self, label=str(self.e_list[0]))

        self.t_ex_40 = wx.StaticText(self, label=expend[0])
        self.t_ex_41 = wx.StaticText(self, label=str(self.e_list[0]))

        self.t_ex_50 = wx.StaticText(self, label=expend[0])
        self.t_ex_51 = wx.StaticText(self, label=str(self.e_list[0]))

        self.t_ex_60 = wx.StaticText(self, label=expend[0])
        self.t_ex_61 = wx.StaticText(self, label=str(self.e_list[0]))

        self.t_ex_70 = wx.StaticText(self, label=expend[0])
        self.t_ex_71 = wx.StaticText(self, label=str(self.e_list[0]))

        self.t_ex_80 = wx.StaticText(self, label=expend[0])
        self.t_ex_81 = wx.StaticText(self, label=str(self.e_list[0]))

        self.t_ex_90 = wx.StaticText(self, label=expend[0])
        self.t_ex_91 = wx.StaticText(self, label=str(self.e_list[0]))

        box3.AddMany([(self.t_ex),(self.t_ex_0,1,wx.EXPAND),
                      (self.t_ex_00),(self.t_ex_01,1,wx.EXPAND),
                      (self.t_ex_10),(self.t_ex_11,1,wx.EXPAND),
                      (self.t_ex_20),(self.t_ex_21,1,wx.EXPAND),
                      (self.t_ex_30),(self.t_ex_31,1,wx.EXPAND),
                      (self.t_ex_40),(self.t_ex_41,1,wx.EXPAND),
                      (self.t_ex_50),(self.t_ex_51,1,wx.EXPAND),
                      (self.t_ex_60),(self.t_ex_61,1,wx.EXPAND),
                      (self.t_ex_70),(self.t_ex_71,1,wx.EXPAND),
                      (self.t_ex_80),(self.t_ex_81,1,wx.EXPAND),
                      (self.t_ex_90),(self.t_ex_91,1,wx.EXPAND)])

        box4 = wx.FlexGridSizer(11, 2, 5, 5)  # 用来展示总支出，每项支出
        self.t_ix = wx.StaticText(self, label=u'总收入')
        self.t_ix_0 = wx.StaticText(self, label=str(self.i_count))

        self.t_ix_00 = wx.StaticText(self, label=imcome[0])
        self.t_ix_01 = wx.StaticText(self, label=str(self.i_list[0]))

        self.t_ix_10 = wx.StaticText(self, label=imcome[1])
        self.t_ix_11 = wx.StaticText(self, label=str(self.i_list[1]))

        self.t_ix_20 = wx.StaticText(self, label=imcome[2])
        self.t_ix_21 = wx.StaticText(self, label=str(self.i_list[2]))

        self.t_ix_30 = wx.StaticText(self, label=imcome[3])
        self.t_ix_31 = wx.StaticText(self, label=str(self.i_list[3]))

        self.t_ix_40 = wx.StaticText(self, label=imcome[4])
        self.t_ix_41 = wx.StaticText(self, label=str(self.i_list[4]))

        self.t_ix_50 = wx.StaticText(self, label=imcome[5])
        self.t_ix_51 = wx.StaticText(self, label=str(self.i_list[5]))

        self.t_ix_60 = wx.StaticText(self, label=imcome[6])
        self.t_ix_61 = wx.StaticText(self, label=str(self.i_list[6]))

        self.t_ix_70 = wx.StaticText(self, label=imcome[7])
        self.t_ix_71 = wx.StaticText(self, label=str(self.i_list[7]))

        self.t_ix_80 = wx.StaticText(self, label=imcome[8])
        self.t_ix_81 = wx.StaticText(self, label=str(self.i_list[8]))

        self.t_ix_90 = wx.StaticText(self, label=imcome[9])
        self.t_ix_91 = wx.StaticText(self, label=str(self.i_list[9]))

        box4.AddMany([(self.t_ix), (self.t_ix_0, 1, wx.EXPAND),
                      (self.t_ix_00), (self.t_ix_01, 1, wx.EXPAND),
                      (self.t_ix_10), (self.t_ix_11, 1, wx.EXPAND),
                      (self.t_ix_20), (self.t_ix_21, 1, wx.EXPAND),
                      (self.t_ix_30), (self.t_ix_31, 1, wx.EXPAND),
                      (self.t_ix_40), (self.t_ix_41, 1, wx.EXPAND),
                      (self.t_ix_50), (self.t_ix_51, 1, wx.EXPAND),
                      (self.t_ix_60), (self.t_ix_61, 1, wx.EXPAND),
                      (self.t_ix_70), (self.t_ix_71, 1, wx.EXPAND),
                      (self.t_ix_80), (self.t_ix_81, 1, wx.EXPAND),
                      (self.t_ix_90), (self.t_ix_91, 1, wx.EXPAND)])
        box5 = wx.BoxSizer(wx.HORIZONTAL)
        box5.Add(box3,flag = wx.LEFT,border = 20)
        box5.Add(box4,flag = wx.LEFT,border = 20)

        box = wx.BoxSizer(wx.VERTICAL)
        box.AddSpacer(30)
        box1 = wx.BoxSizer(wx.HORIZONTAL)
        box1.AddSpacer(30)
        box1.Add(self.ch_year)
        box1.AddSpacer(10)
        box1.Add(self.t_year)
        box1.AddSpacer(10)
        box1.Add(self.ch_month)
        box1.AddSpacer(10)
        box1.Add(self.t_month)
        box.Add(box1)
        box.Add(canvas,0,wx.ALL|wx.EXPAND,0)
        box.Add(box2)
        box.Add(box5)


        self.SetSizer(box)

    def OnChoices1(self,e):
        self.X_year = e.GetString()

    def OnChoices2(self,e):
        self.X_month = e.GetString()

    def GetIncome(self):
        try:
            wb = load_workbook('./data/account.xlsx')
            # print(1)
            ws = wb['收入']
            row = 3
            column = 1
            tody = self.X_year + '/' + self.X_month
            a = ws.cell(row=row, column=column)
            # print(2)
            while a.value:
                a = ws.cell(row=row, column=1)
                time1 = str(ws.cell(row=row, column=1).value)
                #print(time1)
                if tody in time1:
                    kind2 = str(ws.cell(row=row, column=2).value)
                    # print('1')
                    #print(kind2)
                    for i in range(0, 9):
                        #print('2')
                        if kind2 == imcome[i]:
                            #print('3')
                            self.i_list[i] += float(ws.cell(row=row, column=3).value)

                row += 1

        except:
            dlg = wx.MessageDialog(self, '文件打开失败',
                                   '失败',
                                   wx.OK | wx.ICON_INFORMATION
                                   # wx.YES_NO | wx.NO_DEFAULT | wx.CANCEL | wx.ICON_INFORMATION
                                   )
            dlg.ShowModal()
            dlg.Destroy()

        for i in range(0, 9):
            self.i_count += self.i_list[i]

        for i in range(0, 9):
            self.i_space_list[i] = self.i_list[i] / self.i_count

        print(self.i_list)
        print(self.i_space_list)

    def GetExpendture(self):
        try:
            wb = load_workbook('./data/account.xlsx')
            #print(1)
            ws = wb['支出']
            row = 3
            column = 1
            tody = self.X_year + '/' + self.X_month
            a = ws.cell(row=row, column=column)
            #print(2)
            while a.value:
                a = ws.cell(row  =row,column = 1)
                time1=str(ws.cell(row =row,column=1).value)
                #print(time1)
                #print(type(time1))
                print(tody)
                if tody in time1:
                    kind2= str(ws.cell(row =row,column = 2).value)
                    #print('1')
                    #print(kind2)
                    for i in range(0,9):
                        #print('2')
                        if kind2 == expend[i]:
                            print('3')
                            self.e_list[i]+=float(ws.cell(row=row,column=3).value)

                row +=1

        except:
            dlg = wx.MessageDialog(self, '文件打开失败',
                                   '失败',
                                   wx.OK | wx.ICON_INFORMATION
                                   # wx.YES_NO | wx.NO_DEFAULT | wx.CANCEL | wx.ICON_INFORMATION
                                   )
            dlg.ShowModal()
            dlg.Destroy()


        for i in range(0, 9):
            self.e_count += self.e_list[i]

        for i in range(0, 9):
            self.e_space_list[i] = self.e_list[i] / self.e_count

        #print(self.e_list)
        #print(self.e_space_list)