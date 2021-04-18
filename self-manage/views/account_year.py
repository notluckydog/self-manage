import wx
from matplotlib.backends import backend_wxagg
from matplotlib.figure import Figure
import matplotlib
from matplotlib.backends.backend_wxagg import FigureCanvasWxAgg as FigureCanvas
import wx.adv
import time
from openpyxl import load_workbook

imcome = [u'薪资',u'退款',u'转卖',u'兼职',u'还钱',u'借入',u'意外所得',u'报销',u'投资',u'其他']
expend = [u'教育',u'餐饮',u'理财',u'日用',u'零食',u'交通',u'服饰美容',u'数码',u'住房',u'医疗']
year_list = ['2020','2021','2022','2023']
month_list = ['一月','二月','三月','四月','五月','六月','七月','八月','九月','十月','十一月','十二月']


class AccountYear(wx.Panel):
    def __init__(self,parent):
        super().__init__(parent)
        self.time_todey = time.strftime('%Y-%m-%d', time.localtime())
        self.X_year = self.time_todey[0:4]

        self.i_list = [0.1]*10     #用来记录收入列表
        self.e_list = [0.1]*10     #用来记录支出列表
        self.i_space_list = [0.1]*10   #用来记录收入相关占比
        self.e_space_list = [0.1]*10   #用来记录支出相关占比
        self.i_count = 1
        self.e_count = 1
        self.Center()
        self.initUi()

    def initUi(self):
        self.SetBackgroundColour('white')

        # 准备数据
        self.GetIncome()
        self.GetExpendture()
        colors = ['dodgerblue', 'orangered','green','lime','yellow',
                  'limegreen', 'violet', 'gold', 'r','blue',]

        #用来以饼状图的形式展示数据
        self.f = Figure(figsize=(5,6),dpi=100,tight_layout=True)
        matplotlib.rcParams['font.sans-serif'] = ['SimHei']  # 用黑体显示中文
        matplotlib.rcParams['axes.unicode_minus'] = False

        self.sub = self.f.add_subplot(2,2,1,title ="支出明细")

        self.sub.pie(self.e_list, self.e_space_list, expend, colors,
               "%.1f%%", shadow=True, labeldistance=1.1,
                     radius=2,startangle=90)
        self.sub.axis("equal")

        self.sub = self.f.add_subplot(2,2,2,title = "收入明细")
        self.sub.pie(self.i_list, self.i_space_list, imcome, colors,
               "%.1f%%", shadow=True, labeldistance=1.1,
                     radius=2,startangle=90)

        self.sub.axis("equal")

        canvas = FigureCanvas(self, -1, self.f)
        canvas.draw()

        #增加年份、月份选择
        year_list = ['2020', '2021', '2022','2023']
        self.ch_year = wx.Choice(self,-1,(85,18),choices = year_list)
        self.Bind(wx.EVT_CHOICE, self.OnChoices1, self.ch_year)
        self.t_year = wx.StaticText(self,label=u'年')



        #用来展示总支出与每项支出
        box6 = wx.FlexGridSizer(6,4,10,10)
        self.t_ex = wx.StaticText(self,label=u'总支出')
        self.t_ex_0 = wx.StaticText(self,label = str(round(self.e_count,2)))

        self.t_ex_00 = wx.StaticText(self,label =expend[0])
        self.t_ex_01 = wx.StaticText(self,label = str(self.e_list[0]))

        self.t_ex_10 = wx.StaticText(self, label=expend[1])
        self.t_ex_11 = wx.StaticText(self, label=str(self.e_list[1]))

        self.t_ex_20 = wx.StaticText(self, label=expend[2])
        self.t_ex_21 = wx.StaticText(self, label=str(self.e_list[2]))

        self.t_ex_30 = wx.StaticText(self, label=expend[3])
        self.t_ex_31 = wx.StaticText(self, label=str(self.e_list[3]))

        self.t_ex_40 = wx.StaticText(self, label=expend[4])
        self.t_ex_41 = wx.StaticText(self, label=str(self.e_list[4]))

        self.t_ex_50 = wx.StaticText(self, label=expend[5])
        self.t_ex_51 = wx.StaticText(self, label=str(self.e_list[5]))

        self.t_ex_60 = wx.StaticText(self, label=expend[6])
        self.t_ex_61 = wx.StaticText(self, label=str(self.e_list[6]))

        self.t_ex_70 = wx.StaticText(self, label=expend[7])
        self.t_ex_71 = wx.StaticText(self, label=str(self.e_list[7]))

        self.t_ex_80 = wx.StaticText(self, label=expend[8])
        self.t_ex_81 = wx.StaticText(self, label=str(self.e_list[8]))

        self.t_ex_90 = wx.StaticText(self, label=expend[9])
        self.t_ex_91 = wx.StaticText(self, label=str(self.e_list[9]))

        box6.AddMany([(self.t_ex),(self.t_ex_0,1,wx.EXPAND),
                      (wx.StaticText(self)), (wx.StaticText(self)),
                      (self.t_ex_00),(self.t_ex_01,1,wx.EXPAND),
                      (self.t_ex_10),(self.t_ex_11,1,wx.EXPAND),
                      (self.t_ex_20),(self.t_ex_21,1,wx.EXPAND),
                      (self.t_ex_30),(self.t_ex_31,1,wx.EXPAND),
                      (self.t_ex_40),(self.t_ex_41,1,wx.EXPAND),
                      (self.t_ex_50),(self.t_ex_51,1,wx.EXPAND),
                      (self.t_ex_60),(self.t_ex_61,1,wx.EXPAND),
                      (self.t_ex_70),(self.t_ex_71,1,wx.EXPAND),
                      (self.t_ex_80),(self.t_ex_81,1,wx.EXPAND),
                      (self.t_ex_90),(self.t_ex_91,1,wx.EXPAND)])

        box3 = wx.StaticBoxSizer(wx.VERTICAL,self,label = '  ')
        box3.Add(box6)

        #用来展示总收入，每项收入
        box7 = wx.FlexGridSizer(6, 4, 10, 10)
        self.t_ix = wx.StaticText(self, label=u'总收入')
        self.t_ix_0 = wx.StaticText(self, label=str(round(self.i_count,2)))

        self.t_ix_00 = wx.StaticText(self, label=imcome[0])
        self.t_ix_01 = wx.StaticText(self, label=str(self.i_list[0]))

        self.t_ix_10 = wx.StaticText(self, label=imcome[1])
        self.t_ix_11 = wx.StaticText(self, label=str(self.i_list[1]))

        self.t_ix_20 = wx.StaticText(self, label=imcome[2])
        self.t_ix_21 = wx.StaticText(self, label=str(self.i_list[2]))

        self.t_ix_30 = wx.StaticText(self, label=imcome[3])
        self.t_ix_31 = wx.StaticText(self, label=str(self.i_list[3]))

        self.t_ix_40 = wx.StaticText(self, label=imcome[4])
        self.t_ix_41 = wx.StaticText(self, label=str(self.i_list[4]))

        self.t_ix_50 = wx.StaticText(self, label=imcome[5])
        self.t_ix_51 = wx.StaticText(self, label=str(self.i_list[5]))

        self.t_ix_60 = wx.StaticText(self, label=imcome[6])
        self.t_ix_61 = wx.StaticText(self, label=str(self.i_list[6]))

        self.t_ix_70 = wx.StaticText(self, label=imcome[7])
        self.t_ix_71 = wx.StaticText(self, label=str(self.i_list[7]))

        self.t_ix_80 = wx.StaticText(self, label=imcome[8])
        self.t_ix_81 = wx.StaticText(self, label=str(self.i_list[8]))

        self.t_ix_90 = wx.StaticText(self, label=imcome[9])
        self.t_ix_91 = wx.StaticText(self, label=str(self.i_list[9]))

        box7.AddMany([(self.t_ix), (self.t_ix_0, 1, wx.EXPAND),
                      (wx.StaticText(self)),(wx.StaticText(self)),
                      (self.t_ix_00), (self.t_ix_01, 1, wx.EXPAND),
                      (self.t_ix_10), (self.t_ix_11, 1, wx.EXPAND),
                      (self.t_ix_20), (self.t_ix_21, 1, wx.EXPAND),
                      (self.t_ix_30), (self.t_ix_31, 1, wx.EXPAND),
                      (self.t_ix_40), (self.t_ix_41, 1, wx.EXPAND),
                      (self.t_ix_50), (self.t_ix_51, 1, wx.EXPAND),
                      (self.t_ix_60), (self.t_ix_61, 1, wx.EXPAND),
                      (self.t_ix_70), (self.t_ix_71, 1, wx.EXPAND),
                      (self.t_ix_80), (self.t_ix_81, 1, wx.EXPAND),
                      (self.t_ix_90), (self.t_ix_91, 1, wx.EXPAND)])

        box4 = wx.StaticBoxSizer(wx.VERTICAL,self, label =' ')
        box4.Add(box7)

        box5 = wx.BoxSizer(wx.HORIZONTAL)
        box5.Add(box3,flag = wx.LEFT,border = 50)
        box5.AddSpacer(50)
        box5.Add(box4,border = 50)

        box = wx.BoxSizer(wx.VERTICAL)
        box.AddSpacer(30)
        #用来放置日期选择
        box1 = wx.BoxSizer(wx.HORIZONTAL)
        box1.AddSpacer(30)
        box1.Add(self.ch_year)
        box1.AddSpacer(10)
        box1.Add(self.t_year)
        box1.AddSpacer(10)


        box.Add(box1)
        box.Add(box5)
        box.Add(canvas,0,wx.ALL|wx.EXPAND,0)
        self.SetSizer(box)

    def OnChoices1(self,e):
        self.X_year = e.GetString()


    def GetIncome(self):
        try:
            wb = load_workbook('./data/account.xlsx')
            ws = wb['收入']
            row = 3
            column = 1
            tody = self.X_year
            a = ws.cell(row=row, column=column)
            while a.value:
                a = ws.cell(row=row, column=1)
                time1 = str(ws.cell(row=row, column=1).value)
                if tody in time1:
                    kind2 = str(ws.cell(row=row, column=2).value)
                    for i in range(0, 9):
                        if kind2 == imcome[i]:
                            self.i_list[i] += float(ws.cell(row=row, column=3).value)

                row += 1

        except:
            dlg = wx.MessageDialog(self, '文件打开失败',
                                   '失败',
                                   wx.OK | wx.ICON_INFORMATION
                                   # wx.YES_NO | wx.NO_DEFAULT | wx.CANCEL | wx.ICON_INFORMATION
                                   )
            dlg.ShowModal()
            dlg.Destroy()

        for i in range(0, 9):
            self.i_count += self.i_list[i]

        for i in range(0, 9):
            self.i_space_list[i] = self.i_list[i] / self.i_count



    def GetExpendture(self):
        try:
            wb = load_workbook('./data/account.xlsx')
            #print(1)
            ws = wb['支出']
            row = 3
            column = 1
            tody = self.X_year
            a = ws.cell(row=row, column=column)
            while a.value:
                a = ws.cell(row  =row,column = 1)
                time1=str(ws.cell(row =row,column=1).value)
                if tody in time1:
                    kind2= str(ws.cell(row =row,column = 2).value)
                    for i in range(0,9):
                        if kind2 == expend[i]:
                            self.e_list[i]+=float(ws.cell(row=row,column=3).value)

                row +=1

        except:
            dlg = wx.MessageDialog(self, '文件打开失败',
                                   '失败',
                                   wx.OK | wx.ICON_INFORMATION
                                   # wx.YES_NO | wx.NO_DEFAULT | wx.CANCEL | wx.ICON_INFORMATION
                                   )
            dlg.ShowModal()
            dlg.Destroy()


        for i in range(0, 9):
            self.e_count += self.e_list[i]


        for i in range(0, 9):
            self.e_space_list[i] = self.e_list[i] / self.e_count
